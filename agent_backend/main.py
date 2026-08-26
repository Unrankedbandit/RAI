"""Red Flag agent backend — FastAPI surface. The Next.js dashboard calls these
endpoints; the agent pipeline runs as a background task and streams status."""
from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import traceback
import uuid
from pathlib import Path
from typing import Literal

from fastapi import BackgroundTasks, FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, ValidationError

from . import db, grid, memo, redis_state, review, share
from .db import DATABASE_URL
from .agents.base import Agent
from .agents.roles import ANALYST, ROLE_TOOLS
from .gate import GapGate
from .obs import Trace
from .pipeline import _degrade, run_pipeline
from .port_client import PortReporter, port as _port
from .schemas import ChatAnswer, Report
from .telemetry import init_telemetry
from .tools import DOC_DIR

from contextlib import asynccontextmanager


@asynccontextmanager
async def _lifespan(_app: FastAPI):
    # Database + Redis init (no-op when DATABASE_URL / REDIS_URL unset —
    # the existing file/in-memory behavior is the fallback).
    await db.init_pool()
    await db.run_migrations()
    await redis_state.init_client()
    # Crash recovery: pipeline jobs live in process memory, so a restart
    # orphans their Port entities mid-flight. Sweep any RUNNING zombies to
    # FAILED/WorkerLost before serving — the catalog must never claim a run
    # is alive when nothing is executing it.
    _port.reconcile_orphans(set(JOB_QUEUES))
    # Grid warmup: parse+index the grid/blocker GeoJSON in a daemon thread so
    # the first parcel click never pays the ~25 s cold load and the event
    # loop stays free (grid endpoints 503 until loaded — frontend hides it).
    grid.preload()
    yield
    await db.close_pool()
    await redis_state.close_client()


app = FastAPI(title="Red Flag Agent Backend", lifespan=_lifespan)
# Explicit origins + credentials: the hackathon gate authenticates via an
# HttpOnly cookie on .josephbissell.com, so browser fetches from the parcel
# viewer arrive credentialed — and browsers reject allow_origins=["*"] for
# credentialed requests. localhost ports cover local dev.
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://parcel.josephbissell.com",
        "https://rai-parcels.josephbissell.com",
        "https://rai-projects.josephbissell.com",
        "https://mockup.josephbissell.com",
        "https://rai.josephbissell.com",
        "http://localhost:5173",
        "http://localhost:4173",
        "http://localhost:3000",
    ],
    allow_origin_regex=r"https://(.*\.)?josephbissell\.com|http://localhost:\d+",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# SigNoz/OpenTelemetry export. No-op unless SIGNOZ_INGESTION_KEY or
# OTEL_EXPORTER_OTLP_ENDPOINT is set — see agent_backend/telemetry.py.
init_telemetry(app)

# Memo writer (W5) + public share links (W6) — routers live in their own
# modules so this surface stays a thin composition root.
app.include_router(memo.router)
app.include_router(share.router)
app.include_router(grid.router)

# File-based report store — used ONLY when DATABASE_URL is unset (tests/dev).
# When DB is configured, reports/reviews/shares/answers live in PostgreSQL.
STORE = Path(__file__).resolve().parent / "reports"
STORE.mkdir(exist_ok=True)
ARCHIVE_LIST = STORE / "archived.txt"

logger = logging.getLogger(__name__)


def _concise_validation_error(e: ValidationError) -> str:
    """First-error one-liner — the client needs the broken field, not a
    100-line pydantic dump (and never a 500 traceback)."""
    first = e.errors()[0]
    loc = ".".join(str(p) for p in first["loc"]) or "(root)"
    return f"{loc}: {first['msg']} ({e.error_count()} error(s))"


def _archived_ids() -> set[str]:
    """Load archived IDs from archived.txt — only used in file-fallback mode
    (tests/dev without DATABASE_URL). When DB is configured, the
    reports.archived column is the source of truth."""
    try:
        return {
            line.strip()
            for line in ARCHIVE_LIST.read_text(encoding="utf-8").splitlines()
            if line.strip() and not line.startswith("#")
        }
    except FileNotFoundError:
        return set()

JOB_QUEUES: dict[str, asyncio.Queue] = {}
JOB_TRACES: dict[str, Trace] = {}
# --- Run admission control (sized to the deploy hardware) -------------------
# Every /analyze fans out a full pipeline: CPU-heavy doc extraction, parallel
# gap analyzers, LLM streams. Uncapped, N simultaneous runs stack that load on
# the host at once — that coincidence of spikes is what tripped the deploy
# box's PSU on 2026-08-25 (dual RTX 3090 + Threadripper). Cap concurrent runs
# at PIPELINE_MAX_RUNS; extras queue (SSE narrates position) up to
# PIPELINE_MAX_QUEUE, beyond which /analyze answers 429 so clients back off
# instead of piling work onto a saturated box. Defaults sized for the deploy
# host (2 concurrent runs leave CPU/GPU headroom for the desktop + local
# inference stack); override via env on bigger or smaller hardware.
MAX_RUNS = int(os.getenv("PIPELINE_MAX_RUNS", "2"))
MAX_QUEUE = int(os.getenv("PIPELINE_MAX_QUEUE", "4"))
RUN_STATE = {"active": 0, "queued": 0}  # surfaced as /api/health capacity
_RUN_GATE: asyncio.Semaphore | None = None


def _run_gate() -> asyncio.Semaphore:
    """Built lazily so it binds to the running event loop (same reason as the
    LLM semaphore in agents/base.py)."""
    global _RUN_GATE
    if _RUN_GATE is None:
        _RUN_GATE = asyncio.Semaphore(MAX_RUNS)
    return _RUN_GATE

# Per-job gap-review gate handles (GAP_REVIEW=1). Registered at analyze time,
# popped when the job ends; gate.awaiting marks a run actually parked at the
# gap-review pause point — the resume endpoint's 200/409 contract keys off it.
JOB_GATES: dict[str, GapGate] = {}
# Append-only per-job narration log (status strings + trace events) — the
# reconnect-safe source for /api/jobs/{id}/stream; queues stay for back-compat.
JOB_LOGS: dict[str, list] = {}
# Whole-job watchdog. Every agent is individually capped by pipeline.AGENT_TIMEOUT,
# but a run is many serial phases — on a sick bridge each phase can burn its
# full cap, and before this watchdog the SSE stream simply went quiet forever.
# Bounded here so a job ALWAYS ends in a terminal __DONE__/__ERROR__ frame.
PIPELINE_TIMEOUT = int(os.getenv("PIPELINE_TIMEOUT", "2700"))  # 45 min
# In-memory chat answer cache — ONLY used when DB is not configured (tests/dev).
# When DB is enabled, answers persist to PostgreSQL and this dict stays empty.
_ANSWERS_CACHE: dict[str, ChatAnswer] = {}


class AnalyzeRequest(BaseModel):
    name: str
    location: str
    docs: list[str]
    # Per-run pipeline lane. None = fall back to the PIPELINE_MODE env var,
    # so old clients (e2e.py posts only name/location/docs) keep env behavior.
    mode: Literal["fast", "deep"] | None = None


class AskRequest(BaseModel):
    question: str


class ResumeRequest(BaseModel):
    approved: list[str] = []  # gap ids ("gap-1", ...) the human approved


class ReviewDecisionRequest(BaseModel):
    decision: Literal["APPROVED", "REJECTED"]
    reviewer: str = Field(min_length=1, max_length=80)
    rationale: str | None = Field(default=None, max_length=500)
    override: bool = False  # required to change an already-decided review


ALLOWED_UPLOAD = re.compile(r"\.(pdf|xlsx|csv|docx|txt)$", re.IGNORECASE)


@app.post("/api/uploads")
async def uploads(request: Request,
                  files: list[UploadFile] = File(...),
                  x_hax_user: str | None = Header(None)):
    """Receive the actual dossier files (multipart). Saved into the document
    directory the extractors read, so a subsequent /analyze with the returned
    filenames processes the real uploaded bytes. When the DB is configured,
    each file is also extracted and persisted to the documents table with
    IP-based user attribution."""
    client_ip = db.extract_client_ip(request)
    user_id = await db.resolve_user(client_ip, display_name=x_hax_user)
    saved = []
    for f in files:
        name = Path(f.filename or "").name  # strip any client-supplied path
        if not name or not ALLOWED_UPLOAD.search(name):
            continue
        content = await f.read()
        (DOC_DIR / name).write_bytes(content)

        # Persist to documents table with extracted text (no-op without DB).
        ext = Path(name).suffix.lower().lstrip(".")
        file_type = "xlsx" if ext == "xlsx" else ext if ext in ("pdf", "csv", "docx", "txt") else "txt"
        extracted_text = None
        page_count = None
        sheet_count = None
        if file_type == "pdf":
            try:
                from pypdf import PdfReader
                import io
                reader = PdfReader(io.BytesIO(content))
                page_count = len(reader.pages)
                extracted_text = "\n".join(
                    f"--- page {i+1} ---\n{p.extract_text() or ''}"
                    for i, p in enumerate(reader.pages)
                )[:12000]
            except Exception:
                pass
        elif file_type == "xlsx":
            try:
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                sheet_count = len(wb.sheetnames)
                lines = []
                for ws in wb.worksheets:
                    lines.append(f"=== SHEET: {ws.title} ===")
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) if c is not None else "" for c in row]
                        if any(cells):
                            lines.append(" | ".join(cells))
                extracted_text = "\n".join(lines)[:12000]
            except Exception:
                pass

        await db.save_document(
            name,
            file_type=file_type,
            file_size=len(content),
            page_count=page_count,
            sheet_count=sheet_count,
            extracted_text=extracted_text,
            client_ip=client_ip,
            user_id=user_id,
        )
        saved.append(name)
    return {"files": saved}


@app.post("/api/projects/analyze")
async def analyze(req: AnalyzeRequest, request: Request,
                  x_hax_user: str | None = Header(None)):
    from .agents.base import PROVIDER
    # IP-based user identity: every job run is tied to the caller's IP.
    # X-Hax-User (gate SSO) takes precedence as the display name; the IP
    # is always captured as the user identity for audit and attribution.
    client_ip = db.extract_client_ip(request)
    user_id = await db.resolve_user(client_ip, display_name=x_hax_user)
    # The user label that the pipeline stamps on the Report: SSO name when
    # available, else the raw IP so the report is still attributable.
    user_label = x_hax_user or client_ip
    # Admission control: refuse early only when every run slot is taken AND the
    # queue is full — anything less just queues (see work()). Checked before any
    # job state exists so a 429 leaves nothing behind.
    if _run_gate().locked() and RUN_STATE["queued"] >= MAX_QUEUE:
        raise HTTPException(status_code=429,
                            detail=f"pipeline at capacity ({MAX_RUNS} runs active, "
                                   f"{MAX_QUEUE} queued) — retry in a few minutes")
    job_id = uuid.uuid4().hex[:12]
    queue: asyncio.Queue = asyncio.Queue()
    JOB_QUEUES[job_id] = queue
    JOB_LOGS[job_id] = []

    # The one SSE frame that names the lane — the dashboard badges off this.
    # Explicit request mode wins; the env default fills in for old clients.
    mode = req.mode or os.getenv("PIPELINE_MODE", "fast")

    # Persist job metadata to PostgreSQL (survives restart; no-op without DB).
    await db.save_job(job_id, pipeline_mode=mode, user_email=x_hax_user,
                      client_ip=client_ip, user_id=user_id)

    # One trace per job. Its sink pushes structured events onto the same queue
    # the SSE endpoint drains, so the dashboard and the console see identical
    # activity — and JOB_TRACES keeps it replayable after the run ends.
    # The Port reporter rides the same sink: every phase/agent/tool event is
    # mirrored into factory_run / factory_agent_run entities. Fire-and-forget —
    # with no Port credentials configured it is a no-op.
    reporter = PortReporter(job_id, log=lambda m: trace.event("port", m, level="debug"))

    def sink(ev: dict):
        queue.put_nowait(ev)
        JOB_LOGS[job_id].append(ev)
        reporter.handle_event(ev)
        # Dual-write trace events to Redis (cross-process visibility; no-op
        # without REDIS_URL).
        asyncio.ensure_future(redis_state.append_trace(job_id, ev))

    trace = Trace(job_id, sink=sink)
    JOB_TRACES[job_id] = trace
    trace.event(
        "http.request", "POST /api/projects/analyze",
        project=req.name, location=req.location, documents=req.docs,
    )
    # The one SSE frame that names the lane — the dashboard badges off this.
    trace.event("job.mode", f"pipeline mode: {mode}", mode=mode)
    trace.event("job.created", f"job {job_id} queued")
    reporter.start(req.name, req.location, req.docs)

    # Missing credential: keep the analyze→jobId contract (the E2E harness and
    # CI depend on it) and surface the cause as the terminal __ERROR__ frame —
    # an HTTP error here breaks the live-feedback pipe the demo runs on.
    _missing = (PROVIDER == "openai" and not os.getenv("LLM_API_KEY")) or (
        PROVIDER != "openai" and not os.getenv("ANTHROPIC_API_KEY"))
    if _missing:
        _which = "LLM_API_KEY" if PROVIDER == "openai" else "ANTHROPIC_API_KEY"
        _msg = f"{_which} not configured — set it in agent_backend/.env (check /api/health)"
        trace.event("job.error", _msg, level="error")
        JOB_LOGS[job_id].append(f"__ERROR__ {_msg}")
        queue.put_nowait(f"__ERROR__ {_msg}")
        return {"jobId": job_id}

    # Gap-review gate handle for this job. Registered even when GAP_REVIEW is
    # off (the pipeline simply never parks on it) so the resume endpoint can
    # answer 409 instead of 404 for a live, non-awaiting job.
    gate = GapGate()
    JOB_GATES[job_id] = gate

    async def work():
        def status(msg: str):
            queue.put_nowait(msg)
            JOB_LOGS[job_id].append(msg)
        # Admission gate: wait for a run slot here; the SSE stream narrates the
        # wait so the dashboard shows queueing instead of going quiet. The slot
        # is released in the finally below, whatever happens.
        RUN_STATE["queued"] += 1
        status(f"[capacity] queued — {RUN_STATE['queued'] - 1} run(s) ahead "
               f"(PIPELINE_MAX_RUNS={MAX_RUNS})")
        await _run_gate().acquire()
        RUN_STATE["queued"] -= 1
        RUN_STATE["active"] += 1
        # Mirror admission counters to Redis (cross-process; no-op without Redis).
        await redis_state.incr_queued()
        await redis_state.incr_active()
        await db.update_job(job_id, status="running")
        status("[capacity] run slot acquired")
        try:
            report = await asyncio.wait_for(
                run_pipeline(
                    req.name, req.location, req.docs, on_status=status, trace=trace,
                    gap_gate=gate, user=user_label, mode=mode,
                ),
                timeout=PIPELINE_TIMEOUT,
            )
            # Persist report — DB when configured, file fallback for tests/dev.
            if db.is_enabled():
                await db.save_report(
                    job_id, report.model_dump(),
                    name=req.name, location=req.location,
                    pipeline_mode=mode, user_email=x_hax_user,
                    client_ip=client_ip, user_id=user_id,
                )
                # Extract all cited sources into the cited_sources table for
                # the frontend's verified/unverified badges.
                await db.save_cited_sources(job_id, report.model_dump())
                trace.event("job.persisted", f"report saved to database (job {job_id})")
            else:
                path = STORE / f"{job_id}.json"
                path.write_text(report.model_dump_json(indent=2), encoding="utf-8")
                trace.event("job.persisted", f"report written to {path}",
                            bytes=path.stat().st_size)
            await db.update_job(job_id, status="completed", report_id=job_id)
            # Human-in-the-loop gate: the run is not "done" in Port until a
            # person reviews the report there and flips status to APPROVED.
            reporter.awaiting_review(
                report_url=f"{os.getenv('APP_PUBLIC_URL', 'http://localhost:8000')}/api/reports/{job_id}",
                readiness=report.readiness, decision=report.decision, report=report,
            )
            trace.event("job.done", f"job {job_id} complete")
            trace.print_summary()
            JOB_LOGS[job_id].append("__DONE__")
            queue.put_nowait("__DONE__")
        except asyncio.TimeoutError:
            # Watchdog fired: the pipeline was cancelled mid-phase. This is the
            # terminal frame the SSE stream was missing when a run "hung".
            msg = (f"job exceeded PIPELINE_TIMEOUT={PIPELINE_TIMEOUT}s — "
                   "aborted instead of hanging; check the bridge and retry")
            trace.error("job.timeout", msg)
            reporter.failed("PipelineTimeout", msg)
            await db.update_job(job_id, status="timeout", error_message=msg)
            trace.print_summary()
            JOB_LOGS[job_id].append(f"__ERROR__ {msg}")
            queue.put_nowait(f"__ERROR__ {msg}")
        except Exception as e:
            # Previously the traceback was swallowed and only str(e) reached the
            # client — which for AgentDidNotConverge was just an agent name.
            trace.error("job.failed", f"{type(e).__name__}: {e}",
                        traceback=traceback.format_exc()[-2000:])
            reporter.failed(type(e).__name__, str(e))
            await db.update_job(job_id, status="failed",
                                error_message=f"{type(e).__name__}: {e}")
            trace.print_summary()
            JOB_LOGS[job_id].append(f"__ERROR__ {type(e).__name__}: {e}")
            queue.put_nowait(f"__ERROR__ {type(e).__name__}: {e}")
        finally:
            # Job over — the gate handle is dead weight. If the run somehow
            # ended while parked, release any future resume call as a 409.
            gate.awaiting = False
            JOB_GATES.pop(job_id, None)
            RUN_STATE["active"] -= 1
            _run_gate().release()
            # Mirror admission counter decrement to Redis.
            await redis_state.decr_active()

    asyncio.create_task(work())
    return {"jobId": job_id}


@app.get("/api/jobs/{job_id}/trace")
async def job_trace(job_id: str):
    """Full structured trace for a job — every phase, LLM call, and tool call
    with timings. This is the end-to-end view."""
    t = JOB_TRACES.get(job_id)
    if t is None:
        return {"error": "unknown job", "jobId": job_id}
    return {"jobId": job_id, "summary": t.summary(), "events": t.dump()}


@app.get("/api/health")
async def health():
    """Is every external dependency this backend needs actually reachable?"""
    from .agents.base import LLM_BASE_URL, LLM_OPENAI_MODEL, PROVIDER

    t = Trace("health")
    if PROVIDER == "openai":
        llm_configured = bool(os.getenv("LLM_API_KEY"))
        llm_info = {
            "provider": "openai",
            "configured": llm_configured,
            "model": LLM_OPENAI_MODEL,
            "baseUrl": LLM_BASE_URL,
        }
        if not llm_configured:
            t.warn("health.llm", "LLM_API_KEY not set — bridge calls will 401")
    else:
        llm_configured = bool(os.getenv("ANTHROPIC_API_KEY"))
        llm_info = {
            "provider": "anthropic",
            "configured": llm_configured,
            "model": os.getenv("ANTHROPIC_MODEL", "claude-opus-5"),
            "effort": os.getenv("AGENT_EFFORT", "high"),
        }
        if not llm_configured:
            t.warn("health.llm", "ANTHROPIC_API_KEY not set — the agent loop cannot run")
    result = {
        "ok": llm_configured,
        "llm": llm_info,
        "webSearch": {"configured": bool(os.getenv("BRIGHTDATA_API_TOKEN"))},
        "port": {"configured": _port.enabled, "apiBase": _port.api_base if _port.enabled else None},
        "docs": {"dir": os.getenv("DOC_DIR"), "knowledgeBase": os.getenv("KB_DIR")},
        "database": {"configured": db.is_enabled(),
                     "env": os.getenv("APP_ENV", "local"),
                     "url": DATABASE_URL[:20] + "…" if DATABASE_URL else None},
        "redis": {"configured": redis_state.is_enabled()},
        # Run admission control (see analyze): lets the dashboard and smoke
        # distinguish "busy" from "sick".
        "capacity": {"maxRuns": MAX_RUNS, "maxQueue": MAX_QUEUE,
                     "active": RUN_STATE["active"], "queued": RUN_STATE["queued"]},
    }
    t.end()
    return result


@app.get("/api/jobs/{job_id}/stream")
async def stream(job_id: str, from_idx: int = 0):
    """SSE feed of agent activity — powers the demo's live 'agents working' narration.

    Reconnect-safe: the stream is a replay of the append-only JOB_LOGS, not a
    consumed queue, so a client that drops mid-research re-opens with
    ?from_idx=<items already seen> and loses nothing. Multiple viewers OK.
    """
    async def events():
        idx = max(0, from_idx)
        terminal = False
        while not terminal:
            log = JOB_LOGS.get(job_id)
            if log is None:
                yield f"data: {json.dumps({'status': '__ERROR__ unknown job'})}\n\n"
                return
            while idx < len(log):
                msg = log[idx]
                idx += 1
                if isinstance(msg, dict):
                    yield f"data: {json.dumps({'event': msg})}\n\n"
                    continue
                yield f"data: {json.dumps({'status': msg})}\n\n"
                if msg.startswith("__DONE__") or msg.startswith("__ERROR__"):
                    terminal = True
            if not terminal:
                await asyncio.sleep(0.25)
    return StreamingResponse(events(), media_type="text/event-stream")


@app.post("/api/jobs/{job_id}/resume")
async def resume(job_id: str, req: ResumeRequest):
    """Human decision at the gap-review gate (GAP_REVIEW=1): the approved gap
    ids the data scouts should chase. 200 while the job is parked at the gate,
    409 when the job exists but isn't awaiting gap review, 404 for unknown
    jobs. The pipeline emits `gate.resolved` itself when it wakes."""
    gate = JOB_GATES.get(job_id)
    if gate is None or not gate.awaiting:
        if job_id not in JOB_TRACES:
            raise HTTPException(status_code=404, detail=f"unknown job {job_id}")
        raise HTTPException(status_code=409, detail=f"job {job_id} is not awaiting gap review")
    gate.resolve(req.approved)
    return {"ok": True, "mode": "approved"}


@app.get("/api/reports/{job_id}")
async def get_report(job_id: str):
    # DB is the sole source of truth when configured.
    if db.is_enabled():
        report = await db.get_report(job_id)
        if report is not None:
            # Embed cited sources in the report response so the frontend gets
            # them in one fetch — no separate /sources call needed.
            sources = await db.get_cited_sources(job_id)
            if sources:
                report["_cited_sources"] = sources
            return report
        raise HTTPException(
            status_code=404,
            detail=f"no report for job {job_id} — the run may have failed; check /api/jobs/{job_id}/trace",
        )
    # File fallback (tests/dev without DATABASE_URL).
    path = STORE / f"{job_id}.json"
    if not path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"no report for job {job_id} — the run may have failed; check /api/jobs/{job_id}/trace",
        )
    raw = path.read_text(encoding="utf-8")
    try:
        # Re-validate on read: a stored file that no longer conforms (schema
        # tightened, disk corruption, hand-edit) must surface as a concise
        # 422, not a raw blob the frontend mis-renders. Return the VALIDATED
        # model (not raw bytes) so legacy spellings normalize — otherwise the
        # portfolio list (model-driven) and this detail view could disagree
        # (review 2026-08-25).
        report = Report.model_validate_json(raw)
    except ValidationError as e:
        raise HTTPException(
            status_code=422,
            detail=f"stored report {job_id} fails schema validation: "
                   f"{_concise_validation_error(e)}",
        )
    return report.model_dump(mode="json")


@app.get("/api/reports/{report_id}/sources")
async def get_report_sources(report_id: str):
    """Cited sources for a finished report — every URL/reference extracted from
    red_flags, contradictions, acquired_data, and timeline entries. The
    frontend uses this to show verified (has URL) vs unverified (no URL)
    badges on findings."""
    if db.is_enabled():
        if not await db.report_exists(report_id):
            raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
        sources = await db.get_cited_sources(report_id)
        if sources is None:
            return []
        return sources
    # File fallback: extract sources from the JSON on the fly.
    path = STORE / f"{report_id}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
    report = json.loads(path.read_text(encoding="utf-8"))
    # Build a minimal response matching the DB shape.
    sources = []
    for idx, flag in enumerate(report.get("red_flags", [])):
        for src in flag.get("sources", []):
            sources.append({"finding_type": "red_flag", "finding_index": idx,
                            "source_text": src, "source_url": None,
                            "source_label": src[:80], "verified": False})
    return sources


@app.get("/api/reports/{report_id}/review")
async def get_review(report_id: str):
    """Review state for a finished report. A report with no review row defaults
    to AWAITING_REVIEW — the pipeline's terminal state since PR #4."""
    if db.is_enabled():
        exists = await db.report_exists(report_id)
        if not exists:
            raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
        db_review = await db.get_review(report_id)
        if db_review is not None:
            return db_review
        return {"status": review.AWAITING, "reviewedBy": None,
                "reviewedAt": None, "rationale": None}
    # File fallback (tests/dev).
    if not (STORE / f"{report_id}.json").exists():
        raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
    return review.load(STORE, report_id)


@app.post("/api/reports/{report_id}/review")
async def post_review(report_id: str, req: ReviewDecisionRequest,
                      request: Request):
    """Human sign-off on the FINAL report, from inside the app where the gaps
    are shown. Writes to PostgreSQL, mirrors to Port (fire-and-forget), and
    emits a `review.decided` trace event. 409 on an already-decided report
    unless `override: true` is passed explicitly."""
    if db.is_enabled():
        exists = await db.report_exists(report_id)
        if not exists:
            raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
        existing = await db.get_review(report_id) or {"status": review.AWAITING, "reviewedBy": None}
    else:
        if not (STORE / f"{report_id}.json").exists():
            raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
        existing = review.load(STORE, report_id)
    if existing["status"] in review.DECIDED and not req.override:
        raise HTTPException(
            status_code=409,
            detail=(f"report {report_id} already {existing['status']} "
                    f"by {existing.get('reviewedBy')} — pass override=true to change the decision"),
        )
    trace = JOB_TRACES.get(report_id) or Trace(report_id)
    # Write to DB (source of truth) or file fallback.
    if db.is_enabled():
        review_ip = db.extract_client_ip(request)
        await db.save_review(report_id, status=req.decision,
                             reviewed_by=req.reviewer, rationale=req.rationale,
                             client_ip=review_ip)
        result = {"status": req.decision, "reviewedBy": req.reviewer,
                  "reviewedAt": None, "rationale": req.rationale}
        trace.event("review.decided", f"{req.reviewer} {req.decision} report {report_id}",
                    decision=req.decision, reviewer=req.reviewer, report_id=report_id)
        return result
    return review.decide(STORE, report_id, decision=req.decision,
                         reviewer=req.reviewer, rationale=req.rationale,
                         client=_port, trace=trace)


@app.post("/api/reports/{report_id}/ask")
async def ask(report_id: str, req: AskRequest, request: Request):
    """Ask a question about a finished report.

    Returns its own job id immediately and narrates on the existing
    /api/jobs/{id}/stream endpoint, so the Ask rail reuses the transport the
    scan already uses rather than introducing a second streaming shape.
    """
    # Load report — DB when configured, file fallback for tests/dev.
    if db.is_enabled():
        if not await db.report_exists(report_id):
            raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
        report = Report.model_validate(await db.get_report(report_id))
    else:
        path = STORE / f"{report_id}.json"
        if not path.exists():
            raise HTTPException(status_code=404, detail=f"unknown report {report_id}")
        report = Report.model_validate_json(path.read_text(encoding="utf-8"))

    ask_id = uuid.uuid4().hex[:12]
    queue: asyncio.Queue = asyncio.Queue()
    JOB_QUEUES[ask_id] = queue

    # Same one-trace-per-job pattern as analyze: the analyst's steps stream to
    # the Ask rail and export to SigNoz under the same job id.
    trace = Trace(ask_id, sink=lambda ev: queue.put_nowait(ev))
    JOB_TRACES[ask_id] = trace

    async def work():
        def status(msg: str):
            queue.put_nowait(msg)
        try:
            # kb_lookup only — the analyst answers from the finished report,
            # so it needs no web tools in this path. Capped by AGENT_TIMEOUT via
            # _degrade: an uncapped analyst on a sick bridge hung the Ask rail
            # the same way bare pipeline agents hung a run.
            # Deliberately NOT run-gated: one cheap analyst call, already
            # capped by the process-wide LLM semaphore — queuing a question
            # behind a full pipeline run would just feel broken.
            answer = await _degrade(
                Agent(
                    "Analyst", ANALYST, ChatAnswer, ROLE_TOOLS["analyst"], status, trace=trace,
                ).run(req.question, {"report": report.model_dump()}),
                ChatAnswer(
                    answer="The analyst timed out before answering — please try again.",
                    grounded=False,
                ),
                status, "Analyst",
            )
            # Persist answer — DB when configured, in-memory fallback for tests.
            if db.is_enabled():
                ask_ip = db.extract_client_ip(request)
                await db.save_chat_answer(
                    ask_id, report_id=report_id, question=req.question,
                    answer=answer.answer, sources=answer.sources,
                    grounded=answer.grounded, client_ip=ask_ip,
                )
            else:
                _ANSWERS_CACHE[ask_id] = answer
            trace.event("job.done", f"ask {ask_id} complete")
            queue.put_nowait("__DONE__")
        except Exception as e:
            trace.error("job.failed", f"{type(e).__name__}: {e}")
            queue.put_nowait(f"__ERROR__ {e}")
        finally:
            trace.end()

    asyncio.create_task(work())
    return {"jobId": ask_id}


@app.get("/api/asks/{ask_id}")
async def get_answer(ask_id: str):
    # DB is the sole source of truth when configured.
    if db.is_enabled():
        db_answer = await db.get_chat_answer(ask_id)
        if db_answer is not None:
            return db_answer
        raise HTTPException(status_code=404, detail=f"no answer for {ask_id}")
    # In-memory fallback (tests/dev without DATABASE_URL).
    if ask_id not in _ANSWERS_CACHE:
        raise HTTPException(status_code=404, detail=f"no answer for {ask_id}")
    return _ANSWERS_CACHE[ask_id].model_dump()


@app.get("/api/projects")
async def portfolio():
    """Portfolio dashboard: every completed report, worst first.

    Archived ids (reports/archived.txt) are hidden from the listing only —
    their reports stay fetchable by id so permalinks keep working. Review
    sidecar files (*.review.json) are not reports and never list.

    When PostgreSQL is configured, the listing is a single indexed query.
    Without a database, falls back to the file-glob scan below.

    File-fallback response envelope: {"projects": [...rows...],
    "skipped_invalid": [names]}. A stored file that fails schema validation
    is SKIPPED (logged + named in skipped_invalid) rather than 500ing the
    entire board — one corrupt file must not poison the portfolio."""
    if db.is_enabled():
        return await db.list_reports() or []
    # File fallback (tests/dev without DATABASE_URL).
    archived = _archived_ids()
    pairs: list[tuple[str, Report]] = []
    skipped: list[str] = []
    for p in STORE.glob("*.json"):
        if p.stem in archived or p.name.endswith(".review.json"):
            continue
        try:
            pairs.append((p.stem, Report.model_validate_json(p.read_text(encoding="utf-8"))))
        except ValidationError as e:
            logger.warning("portfolio: skipping invalid report file %s — %s",
                           p.name, _concise_validation_error(e))
            skipped.append(p.name)
    pairs.sort(key=lambda t: t[1].readiness)
    return {
        "projects": [
            {"id": pid, "project": r.project, "location": r.location, "readiness": r.readiness,
             "decision": r.decision, "user": r.user,
             "dimensions": [d.model_dump() for d in r.dimensions]}
            for pid, r in pairs
        ],
        "skipped_invalid": sorted(skipped),
    }
