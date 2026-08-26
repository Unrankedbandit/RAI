#!/usr/bin/env python3
"""Seed existing project-docs into the documents table.

Categorizes files as 'positive' (Solar Alpha dossier) or 'negative' (site
comparison contrast documents), extracts text with pypdf/openpyxl, and
persists every document with its extracted content.

Run from repo root:
  PATH=".venv/bin:$PATH" \
  DATABASE_URL="postgresql://rai:rai_dev_pass@localhost:5432/rai" \
  python scripts/seed_documents.py
"""
from __future__ import annotations

import asyncio
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import asyncpg

# Negative-data files: site comparison contrast documents
NEGATIVE_FILES = {
    "08_Site_Comparison_A_Boulder_City_Viable.pdf",
    "09_Site_Comparison_B_Sloan_Canyon_No_Go.pdf",
    "live-boulder-city.pdf",
}

# Positive-data project name / location
POSITIVE_PROJECT = "Solar Alpha"
POSITIVE_LOCATION = "Solano County, California"

NEGATIVE_PROJECT = "RAI Site Comparison"
NEGATIVE_LOCATION = "Southern Nevada"


def extract_pdf(path: Path) -> tuple[str | None, int | None]:
    """Returns (text, page_count)."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        page_count = len(reader.pages)
        text = "\n".join(
            f"--- page {i+1} ---\n{p.extract_text() or ''}"
            for i, p in enumerate(reader.pages)
        )[:12000]
        return text, page_count
    except Exception as e:
        print(f"  WARN: failed to extract {path.name}: {e}")
        return None, None


def extract_xlsx(path: Path) -> tuple[str | None, int | None]:
    """Returns (text, sheet_count)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        sheet_count = len(wb.sheetnames)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"=== SHEET: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        return "\n".join(lines)[:12000], sheet_count
    except Exception as e:
        print(f"  WARN: failed to extract {path.name}: {e}")
        return None, None


async def main() -> None:
    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        print("ERROR: DATABASE_URL not set")
        sys.exit(1)

    docs_dir = Path("project-docs")
    if not docs_dir.exists():
        print(f"ERROR: {docs_dir} not found")
        sys.exit(1)

    conn = await asyncpg.connect(db_url)

    # Clear existing seeded documents (keep any uploaded by users)
    await conn.execute("DELETE FROM documents WHERE job_id IS NULL AND report_id IS NULL")
    print("Cleared previous seeded documents.\n")

    inserted = 0
    for path in sorted(docs_dir.iterdir()):
        if path.name.startswith("."):
            continue
        ext = path.suffix.lower().lstrip(".")
        if ext not in ("pdf", "xlsx", "csv", "docx", "txt"):
            continue

        is_negative = path.name in NEGATIVE_FILES
        category = "negative" if is_negative else "positive"
        project = NEGATIVE_PROJECT if is_negative else POSITIVE_PROJECT
        location = NEGATIVE_LOCATION if is_negative else POSITIVE_LOCATION

        print(f"  Processing: {path.name} [{category}]")

        extracted_text = None
        page_count = None
        sheet_count = None

        if ext == "pdf":
            extracted_text, page_count = extract_pdf(path)
        elif ext == "xlsx":
            extracted_text, sheet_count = extract_xlsx(path)

        file_size = path.stat().st_size

        row = await conn.fetchrow(
            """
            INSERT INTO documents (filename, file_type, category, file_size,
                                   page_count, sheet_count, extracted_text,
                                   text_chars, project_name, location)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
            RETURNING id
            """,
            path.name,
            ext,
            category,
            file_size,
            page_count,
            sheet_count,
            extracted_text,
            len(extracted_text) if extracted_text else 0,
            project,
            location,
        )
        if row:
            inserted += 1
            print(f"    -> id={row['id']} "
                  f"{'(' + str(page_count) + ' pages)' if page_count else ''}"
                  f"{'(' + str(sheet_count) + ' sheets)' if sheet_count else ''}"
                  f" {len(extracted_text or '')} chars extracted")

    await conn.close()
    print(f"\n✅ Seeded {inserted} documents into the database.")

    # Summary
    conn = await asyncpg.connect(db_url)
    stats = await conn.fetch(
        "SELECT category, count(*) as cnt, sum(text_chars) as total_chars "
        "FROM documents WHERE job_id IS NULL GROUP BY category"
    )
    for row in stats:
        print(f"  {row['category']}: {row['cnt']} docs, {row['total_chars'] or 0} chars total")
    await conn.close()


if __name__ == "__main__":
    asyncio.run(main())
